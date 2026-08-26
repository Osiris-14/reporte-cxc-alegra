# -*- coding: utf-8 -*-
"""
Facturas cuya Fecha de apertura == Fecha de vencimiento (mismo dia)
y que registran pagos POSTERIORES a esa fecha en cxc_Pagos.csv.
Fuente: CSV locales del repo.
"""
import csv
from datetime import date
from collections import OrderedDict, defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

AZUL_OSCURO = "1F3864"
AZUL_CLARO = "DDEEFF"
BLANCO = "FFFFFF"
FMT_MONTO = '$#,##0'

MESES = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
         "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]

METODOS = {"cash": "Efectivo", "transfer": "Transferencia", "check": "Cheque",
           "debit-card": "Tarjeta debito", "credit-card": "Tarjeta credito",
           "": "N/D"}


def num(s):
    """Los dos CSV usan formatos distintos: '35752.6' y '30000,0'."""
    s = (s or "").strip().replace('"', '')
    if not s:
        return 0.0
    if ',' in s and '.' in s:
        if s.rfind(',') > s.rfind('.'):
            s = s.replace('.', '').replace(',', '.')
        else:
            s = s.replace(',', '')
    else:
        s = s.replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return 0.0


def fecha(s):
    s = (s or "").strip()[:10]
    try:
        return date.fromisoformat(s)
    except ValueError:
        return None


def leer(path):
    with open(path, encoding='utf-8-sig', newline='') as f:
        return list(csv.DictReader(f))


# ---------------------------------------------------------------- datos
facturas = OrderedDict()
for r in leer('cxc_Cuentasporcobrar.csv'):
    ncf = r['NumeroComprobante'].strip()
    if ncf and ncf not in facturas:
        facturas[ncf] = r

# criterio 1: apertura == vencimiento
mismo_dia = {
    ncf: r for ncf, r in facturas.items()
    if fecha(r['Fecha']) and fecha(r['Fecha']) == fecha(r['FechaVencimiento'])
}

# criterio 2: pagos posteriores al vencimiento
detalle = []
for p in leer('cxc_Pagos.csv'):
    ncf = p['NumeroComprobante'].strip()
    inv = mismo_dia.get(ncf)
    if not inv:
        continue
    fp, fv = fecha(p['FechaPago']), fecha(inv['FechaVencimiento'])
    if not fp or fp <= fv:
        continue
    cliente = (inv['Cliente'] or '').strip() or (p['Cliente'] or '').strip()
    detalle.append({
        'ncf': ncf,
        'cliente': cliente,
        'f_apertura': fv,
        'monto_total': num(inv['MontoTotal']),
        'balance': num(inv['BalancePendiente']),
        'estado': (inv['Estado'] or '').strip(),
        'f_pago': fp,
        'monto_pago': num(p['MontoPago']),
        'metodo': METODOS.get((p['MetodoPago'] or '').strip(), (p['MetodoPago'] or '').strip()),
    })

detalle.sort(key=lambda d: (d['f_apertura'], d['ncf'], d['f_pago']))

# ---------------------------------------------------------------- estilos
fill_head = PatternFill('solid', fgColor=AZUL_OSCURO)
fill_alt = PatternFill('solid', fgColor=AZUL_CLARO)
fill_wht = PatternFill('solid', fgColor=BLANCO)
f_head = Font(name='Arial', size=10, bold=True, color=BLANCO)
f_base = Font(name='Arial', size=10)
f_titulo = Font(name='Arial', size=14, bold=True, color=AZUL_OSCURO)
_s = Side(style='thin', color='B7C9E2')
borde = Border(left=_s, right=_s, top=_s, bottom=_s)

TITULO = "Facturas con pago posterior al vencimiento — 2026"


def encabezar(ws, cols, subtitulo):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
    c = ws.cell(row=1, column=1, value=TITULO)
    c.font = f_titulo
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 24

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(cols))
    c = ws.cell(row=2, column=1, value=subtitulo)
    c.font = Font(name='Arial', size=9, italic=True, color='555555')

    for i, (nombre, _tipo, ancho) in enumerate(cols, start=1):
        c = ws.cell(row=4, column=i, value=nombre)
        c.font, c.fill, c.border = f_head, fill_head, borde
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = ancho
    ws.row_dimensions[4].height = 30
    ws.freeze_panes = 'A5'


def escribir(ws, fila, vals, cols, fill, font, izq=()):
    for j, (v, (_n, tipo, _w)) in enumerate(zip(vals, cols), start=1):
        c = ws.cell(row=fila, column=j, value=v)
        c.font, c.fill, c.border = font, fill, borde
        if tipo == 'mon':
            c.number_format = FMT_MONTO
            c.alignment = Alignment(horizontal='right')
        elif tipo == 'fec':
            c.number_format = 'dd/mm/yyyy'
            c.alignment = Alignment(horizontal='center')
        else:
            c.alignment = Alignment(horizontal='left' if j in izq else 'center')


wb = Workbook()

# ---------------------------------------------------------------- HOJA 1
ws = wb.active
ws.title = "Detalle"
COLS1 = [
    ("NCF", 'txt', 18), ("Cliente", 'txt', 38),
    ("Fecha Apertura/Vencimiento", 'fec', 16), ("Monto Total", 'mon', 15),
    ("Balance Pendiente", 'mon', 15), ("Estado", 'txt', 12),
    ("Fecha Pago Posterior", 'fec', 16), ("Monto Pago Posterior", 'mon', 16),
    ("Método Pago", 'txt', 16),
]
encabezar(ws, COLS1,
          "Una fila por cada pago registrado después de la fecha de vencimiento. "
          "Fuente: cxc_Cuentasporcobrar.csv y cxc_Pagos.csv")

fila = 5
for i, d in enumerate(detalle):
    vals = [d['ncf'], d['cliente'], d['f_apertura'], d['monto_total'], d['balance'],
            d['estado'], d['f_pago'], d['monto_pago'], d['metodo']]
    escribir(ws, fila, vals, COLS1, fill_alt if i % 2 else fill_wht, f_base, izq=(2,))
    fila += 1

# totales hoja 1: monto total y balance se cuentan una sola vez por factura
vistos, tot_monto, tot_bal = set(), 0.0, 0.0
for d in detalle:
    if d['ncf'] not in vistos:
        vistos.add(d['ncf'])
        tot_monto += d['monto_total']
        tot_bal += d['balance']
tot_pagos = sum(d['monto_pago'] for d in detalle)

for j in range(1, len(COLS1) + 1):
    c = ws.cell(row=fila, column=j)
    c.fill, c.font, c.border = fill_head, f_head, borde
ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=3)
c = ws.cell(row=fila, column=1,
            value="TOTALES ({} facturas / {} pagos)".format(len(vistos), len(detalle)))
c.alignment = Alignment(horizontal='left', vertical='center')
for col, val in ((4, tot_monto), (5, tot_bal), (8, tot_pagos)):
    c = ws.cell(row=fila, column=col, value=val)
    c.number_format = FMT_MONTO
    c.alignment = Alignment(horizontal='right')
ws.auto_filter.ref = "A4:{}{}".format(get_column_letter(len(COLS1)), fila - 1)

# ---------------------------------------------------------------- HOJA 2
ws2 = wb.create_sheet("Resumen por mes 2026")
COLS2 = [("Mes", 'txt', 18), ("# Facturas", 'int', 12),
         ("Monto Total aperturado", 'mon', 20),
         ("Monto cobrado posterior", 'mon', 20), ("Balance", 'mon', 16)]
encabezar(ws2, COLS2,
          "Facturas con apertura y vencimiento el mismo día, abiertas en 2026, "
          "agrupadas por mes de apertura")

por_mes = defaultdict(lambda: {'ncf': set(), 'monto': 0.0, 'cobrado': 0.0, 'balance': 0.0})
vistos = set()
for d in detalle:
    if d['f_apertura'].year != 2026:
        continue
    g = por_mes[d['f_apertura'].month]
    if d['ncf'] not in vistos:
        vistos.add(d['ncf'])
        g['ncf'].add(d['ncf'])
        g['monto'] += d['monto_total']
        g['balance'] += d['balance']
    g['cobrado'] += d['monto_pago']

fila = 5
tot = {'n': 0, 'monto': 0.0, 'cobrado': 0.0, 'balance': 0.0}
for i, mes in enumerate(sorted(por_mes)):
    g = por_mes[mes]
    vals = [MESES[mes - 1], len(g['ncf']), g['monto'], g['cobrado'], g['balance']]
    tot['n'] += len(g['ncf'])
    tot['monto'] += g['monto']
    tot['cobrado'] += g['cobrado']
    tot['balance'] += g['balance']
    escribir(ws2, fila, vals, COLS2, fill_alt if i % 2 else fill_wht, f_base, izq=(1,))
    fila += 1

escribir(ws2, fila, ["TOTAL 2026", tot['n'], tot['monto'], tot['cobrado'], tot['balance']],
         COLS2, fill_head, f_head, izq=(1,))

wb.save('facturas_pago_posterior.xlsx')

print("Detalle : {} pagos / {} facturas".format(len(detalle), len({d['ncf'] for d in detalle})))
print("          total ${:,.0f} | balance ${:,.0f} | cobrado posterior ${:,.0f}".format(
    tot_monto, tot_bal, tot_pagos))
print("Resumen : {} meses de 2026 | {} facturas | aperturado ${:,.0f} | cobrado ${:,.0f} | balance ${:,.0f}".format(
    len(por_mes), tot['n'], tot['monto'], tot['cobrado'], tot['balance']))
